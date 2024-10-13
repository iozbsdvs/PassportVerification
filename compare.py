import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def compare_data(df_excel, df_word):
    try:
        logging.info("Начало сравнения данных между Excel и Word.")

        # Создаем новый Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Сравнение данных"

        # Заголовки
        ws.append(['Имя сервера', 'Сайзинг', 'IP адрес', 'Результат'])

        # Цвета для ячеек
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for idx, row_excel in df_excel.iterrows():
            # Убираем пробелы с помощью .strip()
            server_name_excel = row_excel['Имя сервера'].strip()
            sizing_excel = row_excel['Сайзинг'].strip()
            ip_excel = row_excel['IP адрес'].strip()

            logging.info(f"Сравнение ВМ Excel: {server_name_excel} | Сайзинг: {sizing_excel} | IP: {ip_excel}")

            # Поиск совпадений в Word
            row_word = df_word[df_word['Имя сервера'].str.strip() == server_name_excel]  # .str.strip() для Word-данных

            if not row_word.empty:
                name_word = row_word.iloc[0]['Имя сервера'].strip()
                sizing_word = row_word.iloc[0]['Сайзинг'].strip()
                ip_word = row_word.iloc[0]['IP адрес'].strip()

                # Сравнение имени сервера
                name_status = 'Совпадает' if server_name_excel == name_word else 'Не совпадает'

                # Сравнение сайзинга
                sizing_status = 'Совпадает' if sizing_excel == sizing_word else 'Не совпадает'

                # Сравнение IP
                ip_status = 'Совпадает' if ip_excel == ip_word else 'Не совпадает'

                # Логирование результата для ВМ с тремя параметрами
                logging.info(
                    f"ВМ Word: {server_name_excel} | Сайзинг: {sizing_word} | IP: {ip_word} | Результат: {name_status}/{sizing_status}/{ip_status}")
                result = f'{name_status}/{sizing_status}/{ip_status}'

            else:
                logging.info(f"ВМ Word для {server_name_excel} не найдена.")
                name_status = 'Не найдено'
                sizing_status = 'Не найдено'
                ip_status = 'Не найдено'
                result = 'Не найдено/Не найдено/Не найдено'

            # Запись в Excel файл
            row_to_append = [server_name_excel, sizing_excel, ip_excel, result]
            ws.append(row_to_append)

            # Применение цветов
            ws.cell(row=idx + 2, column=1).fill = green_fill if name_status == 'Совпадает' else red_fill
            ws.cell(row=idx + 2, column=2).fill = green_fill if sizing_status == 'Совпадает' else red_fill
            ws.cell(row=idx + 2, column=3).fill = green_fill if ip_status == 'Совпадает' else red_fill

        # Автонастройка ширины колонок
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length

        # Сохранение Excel файла
        wb.save("comparison_results.xlsx")
        logging.info("Результаты сохранены в файл comparison_results.xlsx")
    except Exception as e:
        logging.error(f"Ошибка при сравнении данных: {e}")
